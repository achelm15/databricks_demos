"""
Generates sample_data/salt_river_fields.xlsx with 4 sheets:
  - Attendance
  - Concessions
  - Merchandise
  - Season_Summary
"""

import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)

# ── helpers ──────────────────────────────────────────────────────────────────

def header_row(ws, cols, fill_hex="1B3E6B"):
    fill = PatternFill("solid", fgColor=fill_hex)
    bold_white = Font(bold=True, color="FFFFFF")
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = fill
        cell.font = bold_white
        cell.alignment = Alignment(horizontal="center")

def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

# ── sheet 1: Attendance ──────────────────────────────────────────────────────

ROCKIES_OPP  = ["Padres","Royals","Brewers","Cubs","White Sox","Angels","Astros","Giants","Athletics"]
DBACKS_OPP   = ["Brewers","Cubs","Rockies","Padres","Rangers","Mariners","Dodgers","Reds","Pirates"]
CAPACITY     = 11000

def make_game_schedule():
    games = []
    start = date(2025, 2, 22)
    used = set()
    for i in range(18):
        d = start + timedelta(days=i * 2)
        team = "Colorado Rockies" if i % 2 == 0 else "Arizona Diamondbacks"
        opps = ROCKIES_OPP if team == "Colorado Rockies" else DBACKS_OPP
        opp = opps[i % len(opps)]
        att = random.randint(7200, CAPACITY)
        gate = round(att * random.uniform(18, 28), 2)
        games.append({
            "game_date": d.isoformat(),
            "home_team": team,
            "opponent": opp,
            "attendance": att,
            "capacity": CAPACITY,
            "pct_full": round(att / CAPACITY * 100, 1),
            "gate_revenue_usd": gate,
        })
    return games

def write_attendance(wb, games):
    ws = wb.create_sheet("Attendance")
    cols = ["game_date","home_team","opponent","attendance","capacity","pct_full","gate_revenue_usd"]
    header_row(ws, cols)
    for r, g in enumerate(games, 2):
        for c, k in enumerate(cols, 1):
            ws.cell(row=r, column=c, value=g[k])
    autofit(ws)

# ── sheet 2: Concessions ─────────────────────────────────────────────────────

ITEMS = [
    ("Hot Dog",          5.00,  "Food"),
    ("Bratwurst",        8.50,  "Food"),
    ("Nachos",           9.00,  "Food"),
    ("Soft Pretzel",     7.00,  "Food"),
    ("Peanuts",          4.50,  "Food"),
    ("Cracker Jack",     4.00,  "Food"),
    ("Cotton Candy",     5.50,  "Food"),
    ("Budweiser",       12.00,  "Beer"),
    ("Coors Light",     12.00,  "Beer"),
    ("Arizona Ale",     14.00,  "Beer"),
    ("Hard Seltzer",    13.00,  "Beer"),
    ("Lemonade",         6.00,  "Non-Alcoholic"),
    ("Soda (Pepsi)",     5.50,  "Non-Alcoholic"),
    ("Water",            4.00,  "Non-Alcoholic"),
    ("Loaded Fries",    10.00,  "Food"),
]

def write_concessions(wb, games):
    ws = wb.create_sheet("Concessions")
    cols = ["game_date","home_team","item_name","category","unit_price_usd","units_sold","total_revenue_usd"]
    header_row(ws, cols, fill_hex="8B1A1A")
    row = 2
    for g in games:
        for item, price, cat in ITEMS:
            units = random.randint(30, 450)
            ws.cell(row=row, column=1, value=g["game_date"])
            ws.cell(row=row, column=2, value=g["home_team"])
            ws.cell(row=row, column=3, value=item)
            ws.cell(row=row, column=4, value=cat)
            ws.cell(row=row, column=5, value=price)
            ws.cell(row=row, column=6, value=units)
            ws.cell(row=row, column=7, value=round(units * price, 2))
            row += 1
    autofit(ws)

# ── sheet 3: Merchandise ─────────────────────────────────────────────────────

MERCH = [
    ("COL-HAT-ADJ",    "Rockies Adjustable Cap",        "Headwear",   35.00,  "Rockies"),
    ("COL-HAT-59FIFTY","Rockies 59FIFTY Fitted",         "Headwear",   45.00,  "Rockies"),
    ("COL-JSY-HOME",   "Rockies Home Replica Jersey",   "Apparel",   130.00,  "Rockies"),
    ("COL-JSY-ROAD",   "Rockies Road Replica Jersey",   "Apparel",   130.00,  "Rockies"),
    ("COL-TEE-LOGO",   "Rockies Logo Tee",              "Apparel",    32.00,  "Rockies"),
    ("COL-HOO-PULL",   "Rockies Pullover Hoodie",       "Apparel",    75.00,  "Rockies"),
    ("COL-MUG",        "Rockies Souvenir Mug",          "Souvenirs",  18.00,  "Rockies"),
    ("COL-PROG",       "Rockies Spring Training Program","Souvenirs",  10.00,  "Rockies"),
    ("ARI-HAT-ADJ",    "D-backs Adjustable Cap",        "Headwear",   35.00,  "D-backs"),
    ("ARI-HAT-59FIFTY","D-backs 59FIFTY Fitted",         "Headwear",   45.00,  "D-backs"),
    ("ARI-JSY-HOME",   "D-backs Home Replica Jersey",   "Apparel",   130.00,  "D-backs"),
    ("ARI-JSY-ROAD",   "D-backs Road Replica Jersey",   "Apparel",   130.00,  "D-backs"),
    ("ARI-TEE-LOGO",   "D-backs Logo Tee",              "Apparel",    32.00,  "D-backs"),
    ("ARI-HOO-PULL",   "D-backs Pullover Hoodie",       "Apparel",    75.00,  "D-backs"),
    ("ARI-MUG",        "D-backs Souvenir Mug",          "Souvenirs",  18.00,  "D-backs"),
    ("ARI-PROG",       "D-backs Spring Training Program","Souvenirs", 10.00,  "D-backs"),
    ("SRF-HAT",        "Salt River Fields Hat",          "Headwear",   35.00,  "Venue"),
    ("SRF-TEE",        "Salt River Fields Tee",          "Apparel",    30.00,  "Venue"),
    ("SRF-PIN",        "SRF Commemorative Pin",          "Souvenirs",   8.00,  "Venue"),
]

def write_merchandise(wb, games):
    ws = wb.create_sheet("Merchandise")
    cols = ["game_date","home_team","sku","item_name","category","team","unit_price_usd","units_sold","total_revenue_usd"]
    header_row(ws, cols, fill_hex="1B6B2E")
    row = 2
    for g in games:
        for sku, name, cat, price, team in MERCH:
            units = random.randint(0, 60)
            ws.cell(row=row, column=1, value=g["game_date"])
            ws.cell(row=row, column=2, value=g["home_team"])
            ws.cell(row=row, column=3, value=sku)
            ws.cell(row=row, column=4, value=name)
            ws.cell(row=row, column=5, value=cat)
            ws.cell(row=row, column=6, value=team)
            ws.cell(row=row, column=7, value=price)
            ws.cell(row=row, column=8, value=units)
            ws.cell(row=row, column=9, value=round(units * price, 2))
            row += 1
    autofit(ws)

# ── sheet 4: Season_Summary ──────────────────────────────────────────────────

def write_summary(wb, games):
    ws = wb.create_sheet("Season_Summary")
    cols = ["metric","value"]
    header_row(ws, cols, fill_hex="4A235A")
    total_att = sum(g["attendance"] for g in games)
    total_gate = sum(g["gate_revenue_usd"] for g in games)
    metrics = [
        ("Total Games",          len(games)),
        ("Total Attendance",     total_att),
        ("Avg Attendance",       round(total_att / len(games), 0)),
        ("Avg % Full",           round(sum(g["pct_full"] for g in games) / len(games), 1)),
        ("Total Gate Revenue",   round(total_gate, 2)),
        ("Rockies Home Games",   sum(1 for g in games if "Rockies" in g["home_team"])),
        ("D-backs Home Games",   sum(1 for g in games if "Arizona" in g["home_team"])),
        ("Season",               "2025 Cactus League"),
        ("Venue",                "Salt River Fields at Talking Stick"),
        ("Location",             "Scottsdale, AZ"),
    ]
    for r, (m, v) in enumerate(metrics, 2):
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=v)
    autofit(ws)

# ── main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import os
    out = os.path.join(os.path.dirname(__file__), "salt_river_fields.xlsx")
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    games = make_game_schedule()
    write_attendance(wb, games)
    write_concessions(wb, games)
    write_merchandise(wb, games)
    write_summary(wb, games)

    wb.save(out)
    print(f"Saved: {out}")
    for ws in wb.worksheets:
        print(f"  Sheet '{ws.title}': {ws.max_row - 1} data rows, {ws.max_column} cols")
